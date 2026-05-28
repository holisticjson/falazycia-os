"""
📄 Konwerter XLSX/Google Sheets → MD
Użycie: python xlsx_to_md.py plik.xlsx
        python xlsx_to_md.py folder/  (konwertuje wszystkie xlsx w folderze)
"""
import sys
import openpyxl
from pathlib import Path

def xlsx_to_md(filepath):
    filepath = Path(filepath)
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    md_parts = [f"# {filepath.stem}\n"]
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        
        md_parts.append(f"\n## {sheet_name}\n")
        
        # Filtruj puste wiersze
        rows = [r for r in rows if any(cell is not None for cell in r)]
        if not rows:
            continue
        
        # Nagłówek (pierwszy wiersz)
        headers = [str(c) if c else "" for c in rows[0]]
        md_parts.append("| " + " | ".join(headers) + " |")
        md_parts.append("| " + " | ".join(["---"] * len(headers)) + " |")
        
        # Dane
        for row in rows[1:]:
            cells = [str(c).replace("|", "\\|").replace("\n", " ") if c else "" for c in row]
            # Dopasuj liczbę kolumn
            while len(cells) < len(headers):
                cells.append("")
            md_parts.append("| " + " | ".join(cells[:len(headers)]) + " |")
    
    wb.close()
    
    # Zapisz MD
    out_path = filepath.with_suffix(".md")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(md_parts))
    
    print(f"OK: {filepath.name} -> {out_path.name} ({len(rows)} wierszy)")
    return out_path

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Użycie: python xlsx_to_md.py plik.xlsx")
        print("        python xlsx_to_md.py folder/")
        sys.exit(1)
    
    target = Path(sys.argv[1])
    
    if target.is_dir():
        files = list(target.glob("*.xlsx"))
        print(f"📂 Znaleziono {len(files)} plików xlsx w {target}")
        for f in files:
            try:
                xlsx_to_md(f)
            except Exception as e:
                print(f"❌ {f.name}: {e}")
    else:
        xlsx_to_md(target)
